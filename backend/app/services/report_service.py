from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional
from uuid import UUID
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import func, and_

from app.models.business import Customer, Supplier, Company
from app.models.product import Product, ProductCategory
from app.models.purchase import PurchaseOrder, PurchaseEntry
from app.models.inventory import CurrentStock, StockTransaction
from app.models.sales import SalesOrder, Invoice, InvoiceItem
from app.models.finance import Payment, VendorPayment
from app.schemas.transaction import (
    DashboardResponse, KPICardsOut, SalesByCategoryOut,
    MonthlySalesTrendOut, TopProductSalesOut, RecentTransactionOut,
    LedgerEntry, CustomerLedgerResponse, SupplierLedgerResponse
)
from app.schemas.accounts import (
    TrialBalanceResponse,
    GeneralLedgerResponse,
    DayBookResponse,
    PurchaseRegisterResponse,
    SalesRegisterResponse
)


class ReportService:
    @staticmethod
    async def get_dashboard_metrics(db: AsyncSession, company_id: Optional[UUID] = None) -> DashboardResponse:
        """Aggregate data to populate the primary KPI dashboard charts and tables."""
        today = date.today()
        start_of_today = datetime(today.year, today.month, today.day)
        start_of_month = datetime(today.year, today.month, 1)

        # ---------------------------------------------
        # 1. KPI CARDS SUMS
        # ---------------------------------------------
        # Today's Sales
        stmt_today = select(func.sum(Invoice.total_amount)).filter(Invoice.date >= start_of_today)
        if company_id:
            stmt_today = stmt_today.filter(Invoice.company_id == company_id)
        q_today = await db.execute(stmt_today)
        today_sales = q_today.scalar() or 0.0

        # Monthly Sales
        stmt_month = select(func.sum(Invoice.total_amount)).filter(Invoice.date >= start_of_month)
        if company_id:
            stmt_month = stmt_month.filter(Invoice.company_id == company_id)
        q_month = await db.execute(stmt_month)
        monthly_sales = q_month.scalar() or 0.0

        # Outstanding Payments
        stmt_out = select(func.sum(Invoice.total_amount)).filter(Invoice.status != "Paid")
        if company_id:
            stmt_out = stmt_out.filter(Invoice.company_id == company_id)
        q_out = await db.execute(stmt_out)
        outstanding = q_out.scalar() or 0.0

        # Low Stock count (globally common inventory)
        # Select products where min_stock_level > current_stock
        stmt_low = select(func.count(Product.id)).outerjoin(
            CurrentStock, CurrentStock.product_id == Product.id
        ).filter(Product.min_stock_level > func.coalesce(CurrentStock.qty, 0.0))
        q_low = await db.execute(stmt_low)
        low_stock = q_low.scalar() or 0

        kpis = KPICardsOut(
            today_sales=today_sales,
            monthly_sales=monthly_sales,
            outstanding_payments=outstanding,
            low_stock_count=low_stock
        )

        # ---------------------------------------------
        # 2. SALES BY CATEGORY PIE CHART
        # ---------------------------------------------
        stmt_cat = (
            select(ProductCategory.name, func.sum(InvoiceItem.amount))
            .join(Product, Product.category_id == ProductCategory.id)
            .join(InvoiceItem, InvoiceItem.product_id == Product.id)
            .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        )
        if company_id:
            stmt_cat = stmt_cat.filter(Invoice.company_id == company_id)
        stmt_cat = stmt_cat.group_by(ProductCategory.name)
        q_cat = await db.execute(stmt_cat)
        
        sales_by_category = [
            SalesByCategoryOut(category_name=row[0], total_sales=row[1] or 0.0)
            for row in q_cat.all()
        ]

        # ---------------------------------------------
        # 3. MONTHLY SALES TREND (LAST 12 MONTHS)
        # ---------------------------------------------
        monthly_sales_trend = []
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        # Pull invoice sums grouped by month
        for idx, m_name in enumerate(months):
            stmt_trend = select(func.sum(Invoice.total_amount)).filter(func.extract('month', Invoice.date) == (idx + 1))
            if company_id:
                stmt_trend = stmt_trend.filter(Invoice.company_id == company_id)
            q_trend = await db.execute(stmt_trend)
            monthly_sales_trend.append(
                MonthlySalesTrendOut(month=m_name, sales=q_trend.scalar() or 0.0)
            )

        # ---------------------------------------------
        # 4. TOP 10 PRODUCTS BY SALES VOLUME
        # ---------------------------------------------
        stmt_top = (
            select(Product.name, Product.sku, func.sum(InvoiceItem.qty), func.sum(InvoiceItem.amount))
            .join(InvoiceItem, InvoiceItem.product_id == Product.id)
            .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        )
        if company_id:
            stmt_top = stmt_top.filter(Invoice.company_id == company_id)
        stmt_top = stmt_top.group_by(Product.name, Product.sku).order_by(func.sum(InvoiceItem.amount).desc()).limit(10)
        q_top = await db.execute(stmt_top)
        
        top_products = [
            TopProductSalesOut(product_name=row[0], sku=row[1], qty_sold=row[2] or 0.0, total_revenue=row[3] or 0.0)
            for row in q_top.all()
        ]

        # ---------------------------------------------
        # 5. RECENT TRANSACTIONS LOG
        # ---------------------------------------------
        # Fetch last 5 Sales Orders
        stmt_so = select(SalesOrder).options(selectinload(SalesOrder.customer)).order_by(SalesOrder.date.desc()).limit(5)
        if company_id:
            stmt_so = stmt_so.filter(SalesOrder.company_id == company_id)
        q_so = await db.execute(stmt_so)
        
        # Fetch last 5 Purchase Orders
        stmt_po = select(PurchaseOrder).options(selectinload(PurchaseOrder.supplier)).order_by(PurchaseOrder.date.desc()).limit(5)
        if company_id:
            stmt_po = stmt_po.filter(PurchaseOrder.company_id == company_id)
        q_po = await db.execute(stmt_po)

        recent_txs = []
        for so in q_so.scalars().all():
            recent_txs.append(
                RecentTransactionOut(
                    id=so.id,
                    tx_type="Sales Order",
                    reference_no=f"SO-{so.id.hex[:6].upper()}",
                    party_name=so.customer.name,
                    date=so.date,
                    amount=so.grand_total,
                    status=so.status
                )
            )
        
        for po in q_po.scalars().all():
            recent_txs.append(
                RecentTransactionOut(
                    id=po.id,
                    tx_type="Purchase Order",
                    reference_no=f"PO-{po.id.hex[:6].upper()}",
                    party_name=po.supplier.name,
                    date=po.date,
                    amount=po.grand_total,
                    status=po.status
                )
            )
        
        # Sort combined list by date desc
        recent_txs.sort(key=lambda x: x.date, reverse=True)
        recent_transactions = recent_txs[:10]

        return DashboardResponse(
            kpis=kpis,
            sales_by_category=sales_by_category,
            monthly_sales_trend=monthly_sales_trend,
            top_products=top_products,
            recent_transactions=recent_transactions
        )

    @staticmethod
    async def get_customer_ledger(
        db: AsyncSession,
        customer_id: UUID,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        company_id: Optional[UUID] = None
    ) -> CustomerLedgerResponse:
        # 1. Parse start and end datetimes
        start_dt = None
        end_dt = None
        if start_date:
            try:
                if len(start_date) == 7:
                    start_dt = datetime.strptime(start_date, "%Y-%m")
                else:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            except Exception:
                pass
        if end_date:
            try:
                if len(end_date) == 7:
                    year, month = map(int, end_date.split("-"))
                    if month == 12:
                        end_dt = datetime(year + 1, 1, 1) - timedelta(seconds=1)
                    else:
                        end_dt = datetime(year, month + 1, 1) - timedelta(seconds=1)
                else:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            except Exception:
                pass

        # 2. Fetch Invoices
        stmt_inv = (
            select(Invoice)
            .join(SalesOrder, Invoice.sales_order_id == SalesOrder.id)
            .filter(SalesOrder.customer_id == customer_id, Invoice.status != "Cancelled")
        )
        if company_id:
            stmt_inv = stmt_inv.filter(Invoice.company_id == company_id)
        if start_dt:
            stmt_inv = stmt_inv.filter(Invoice.date >= start_dt)
        if end_dt:
            stmt_inv = stmt_inv.filter(Invoice.date <= end_dt)
        q_inv = await db.execute(stmt_inv.order_by(Invoice.date.asc()))
        invoices = q_inv.scalars().all()

        # 3. Fetch Payments
        stmt_pay = select(Payment).filter(Payment.customer_id == customer_id)
        if company_id:
            stmt_pay = stmt_pay.join(Invoice, Payment.invoice_id == Invoice.id, isouter=True).filter(
                (Invoice.company_id == company_id) | (Payment.invoice_id == None)
            )
        if start_dt:
            stmt_pay = stmt_pay.filter(Payment.payment_date >= start_dt)
        if end_dt:
            stmt_pay = stmt_pay.filter(Payment.payment_date <= end_dt)
        q_pay = await db.execute(stmt_pay.order_by(Payment.payment_date.asc()))
        payments = q_pay.scalars().all()

        # 4. Merge entries chronologically
        entries = []
        for inv in invoices:
            entries.append({
                "date": inv.date,
                "tx_type": "Invoice",
                "reference_no": inv.invoice_number,
                "debit": inv.total_amount,
                "credit": 0.0
            })
        for pay in payments:
            entries.append({
                "date": pay.payment_date,
                "tx_type": "Payment",
                "reference_no": pay.reference_number or "N/A",
                "debit": 0.0,
                "credit": pay.amount_paid
            })

        # Sort entries by date
        entries.sort(key=lambda x: x["date"])

        # Fetch customer opening balance details
        q_cust = await db.execute(select(Customer).filter(Customer.id == customer_id))
        customer = q_cust.scalar_one_or_none()
        opening_bal = customer.opening_bal if customer else 0.0
        opening_bal_type = customer.opening_bal_type if customer else "Dr"

        # Calculate running balance and totals
        total_billed = 0.0
        total_paid = 0.0
        running_balance = opening_bal if opening_bal_type == "Dr" else -opening_bal
        
        ledger_entries = []
        ledger_entries.append(
            LedgerEntry(
                date=start_dt or (customer.created_at if customer else datetime(2000, 1, 1)),
                tx_type="Opening Balance",
                reference_no="OB",
                debit=opening_bal if opening_bal_type == "Dr" else 0.0,
                credit=opening_bal if opening_bal_type == "Cr" else 0.0,
                running_balance=running_balance
            )
        )

        for e in entries:
            debit = e["debit"]
            credit = e["credit"]
            total_billed += debit
            total_paid += credit
            running_balance += (debit - credit)
            ledger_entries.append(
                LedgerEntry(
                    date=e["date"],
                    tx_type=e["tx_type"],
                    reference_no=e["reference_no"],
                    debit=debit,
                    credit=credit,
                    running_balance=running_balance
                )
            )

        return CustomerLedgerResponse(
            total_billed=total_billed,
            total_paid=total_paid,
            balance=running_balance,
            transactions=ledger_entries
        )

    @staticmethod
    async def get_supplier_ledger(
        db: AsyncSession,
        supplier_id: UUID,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        company_id: Optional[UUID] = None
    ) -> SupplierLedgerResponse:
        # 1. Parse start and end datetimes
        start_dt = None
        end_dt = None
        if start_date:
            try:
                if len(start_date) == 7:
                    start_dt = datetime.strptime(start_date, "%Y-%m")
                else:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            except Exception:
                pass
        if end_date:
            try:
                if len(end_date) == 7:
                    year, month = map(int, end_date.split("-"))
                    if month == 12:
                        end_dt = datetime(year + 1, 1, 1) - timedelta(seconds=1)
                    else:
                        end_dt = datetime(year, month + 1, 1) - timedelta(seconds=1)
                else:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            except Exception:
                pass

        # 2. Fetch Bills (PurchaseEntry)
        stmt_bills = select(PurchaseEntry).filter(PurchaseEntry.supplier_id == supplier_id, PurchaseEntry.status != "Cancelled")
        if company_id:
            stmt_bills = stmt_bills.filter(PurchaseEntry.company_id == company_id)
        if start_dt:
            stmt_bills = stmt_bills.filter(PurchaseEntry.billing_date >= start_dt)
        if end_dt:
            stmt_bills = stmt_bills.filter(PurchaseEntry.billing_date <= end_dt)
        q_bills = await db.execute(stmt_bills.order_by(PurchaseEntry.billing_date.asc()))
        bills = q_bills.scalars().all()

        # 3. Fetch Payments (VendorPayment)
        stmt_payments = select(VendorPayment).filter(VendorPayment.supplier_id == supplier_id).options(selectinload(VendorPayment.purchase_entry))
        if company_id:
            stmt_payments = stmt_payments.join(PurchaseEntry, VendorPayment.purchase_entry_id == PurchaseEntry.id, isouter=True).filter(
                (PurchaseEntry.company_id == company_id) | (VendorPayment.purchase_entry_id == None)
            )
        if start_dt:
            stmt_payments = stmt_payments.filter(VendorPayment.payment_date >= start_dt)
        if end_dt:
            stmt_payments = stmt_payments.filter(VendorPayment.payment_date <= end_dt)
        q_payments = await db.execute(stmt_payments.order_by(VendorPayment.payment_date.asc()))
        payments = q_payments.scalars().all()

        # 4. Merge entries chronologically
        entries = []
        for bill in bills:
            entries.append({
                "date": bill.billing_date,
                "tx_type": "Purchase Bill",
                "reference_no": bill.invoice_number,
                "debit": bill.total_amount,
                "credit": 0.0
            })
        for pay in payments:
            ref_no = pay.reference_number
            if not ref_no or ref_no == "-":
                if pay.purchase_entry:
                    ref_no = pay.purchase_entry.invoice_number
                else:
                    ref_no = "Advance Payment"
            elif pay.purchase_entry:
                ref_no = f"{ref_no} ({pay.purchase_entry.invoice_number})"
            
            entries.append({
                "date": pay.payment_date,
                "tx_type": "Payment",
                "reference_no": ref_no,
                "debit": 0.0,
                "credit": pay.amount_paid
            })

        # Sort entries by date
        entries.sort(key=lambda x: x["date"])

        # Fetch supplier opening balance details
        q_supp = await db.execute(select(Supplier).filter(Supplier.id == supplier_id))
        supplier = q_supp.scalar_one_or_none()
        opening_bal = supplier.opening_bal if supplier else 0.0
        opening_bal_type = supplier.opening_bal_type if supplier else "Cr"

        total_purchased = 0.0
        total_paid = 0.0
        running_balance = opening_bal if opening_bal_type == "Cr" else -opening_bal
        
        ledger_entries = []
        ledger_entries.append(
            LedgerEntry(
                date=start_dt or (supplier.created_at if supplier else datetime(2000, 1, 1)),
                tx_type="Opening Balance",
                reference_no="OB",
                debit=opening_bal if opening_bal_type == "Dr" else 0.0,
                credit=opening_bal if opening_bal_type == "Cr" else 0.0,
                running_balance=running_balance
            )
        )

        for e in entries:
            debit = e["debit"]
            credit = e["credit"]
            total_purchased += debit
            total_paid += credit
            running_balance += (debit - credit)
            ledger_entries.append(
                LedgerEntry(
                    date=e["date"],
                    tx_type=e["tx_type"],
                    reference_no=e["reference_no"],
                    debit=debit,
                    credit=credit,
                    running_balance=running_balance
                )
            )

        return SupplierLedgerResponse(
            total_purchased=total_purchased,
            total_paid=total_paid,
            balance=running_balance,
            transactions=ledger_entries
        )

    @staticmethod
    async def get_sales_summary_data(
        db: AsyncSession,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        company_id: Optional[UUID] = None
    ) -> List[Dict[str, Any]]:
        """Fetch and format sales summary item level data for Excel / UI report."""
        stmt = (
            select(Invoice)
            .options(
                selectinload(Invoice.items).selectinload(InvoiceItem.product),
                selectinload(Invoice.sales_order).selectinload(SalesOrder.customer),
                selectinload(Invoice.payments),
                selectinload(Invoice.delivery_challan)
            )
            .filter(Invoice.status != "Cancelled")
        )
        
        if company_id:
            stmt = stmt.filter(Invoice.company_id == company_id)
            
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                stmt = stmt.filter(Invoice.date >= start_dt)
            except ValueError:
                pass
                
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
                stmt = stmt.filter(Invoice.date <= end_dt)
            except ValueError:
                pass
                
        query = await db.execute(stmt.order_by(Invoice.invoice_number.asc()))
        invoices = list(query.scalars().all())
        
        rows = []
        STATE_CODES = {
            "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh", "05": "Uttarakhand",
            "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar",
            "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
            "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
            "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat", "26": "Dadra and Nagar Haveli and Daman and Diu",
            "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
            "38": "Ladakh"
        }

        for inv in invoices:
            # Gather company state details
            company_state = "22"
            if inv.company_id:
                q_comp = await db.execute(select(Company).filter(Company.id == inv.company_id))
                company = q_comp.scalar_one_or_none()
                if company:
                    company_state = company.gstin[:2] if (company.gstin and len(company.gstin) >= 2) else (company.state_code if company.state_code else "22")
            
            # Aggregate payment details
            total_paid = sum([p.amount_paid for p in inv.payments])
            outstanding_amount = max(0.0, inv.total_amount - total_paid)
            
            payment_date = ""
            payment_mode = ""
            if inv.payments:
                sorted_payments = sorted(inv.payments, key=lambda x: x.payment_date, reverse=True)
                payment_date = sorted_payments[0].payment_date.strftime("%Y-%m-%d")
                payment_mode = sorted_payments[0].payment_mode.upper()
                
            # Customer details
            cust_name = "Unknown"
            cust_gstin = ""
            cust_address = ""
            customer = None
            if inv.sales_order and inv.sales_order.customer:
                customer = inv.sales_order.customer
            elif inv.delivery_challan and inv.delivery_challan.customer:
                customer = inv.delivery_challan.customer
                
            if customer:
                cust_name = customer.name
                cust_gstin = customer.gstin or ""
                cust_address = customer.billing_address or ""
                
            # Place of supply
            place_of_supply = "Other"
            cust_state_code = cust_gstin[:2] if (cust_gstin and len(cust_gstin) >= 2) else None
            if cust_state_code in STATE_CODES:
                place_of_supply = f"{STATE_CODES[cust_state_code]} ({cust_state_code})"
            elif cust_address:
                place_of_supply = cust_address.split(",")[-1].strip()
                
            is_interstate = (cust_state_code is not None) and (cust_state_code != company_state)
            
            # Iterate invoice items
            for item in inv.items:
                product_name = item.product.name if item.product else "Unknown Product"
                hsn_code = item.product.hsn_code if item.product else ""
                
                taxable_value = item.amount
                discount = item.discount_amount
                
                # GST Breakdown for this item
                cgst_pct = 0.0
                cgst_amt = 0.0
                sgst_pct = 0.0
                sgst_amt = 0.0
                igst_pct = 0.0
                igst_amt = 0.0
                
                if is_interstate:
                    igst_pct = item.tax_rate
                    igst_amt = item.tax_amount
                else:
                    cgst_pct = item.tax_rate / 2.0
                    cgst_amt = item.tax_amount / 2.0
                    sgst_pct = item.tax_rate / 2.0
                    sgst_amt = item.tax_amount / 2.0
                    
                rows.append({
                    "id": str(item.id),
                    "invoice_id": str(inv.id),
                    "invoice_number": inv.invoice_number,
                    "invoice_date": inv.date.strftime("%Y-%m-%d"),
                    "payment_date": payment_date,
                    "payment_mode": payment_mode,
                    "customer_name": cust_name,
                    "customer_gstin": cust_gstin,
                    "place_of_supply": place_of_supply,
                    "hsn_code": hsn_code,
                    "product_description": product_name,
                    "taxable_value": round(taxable_value, 2),
                    "discount": round(discount, 2),
                    "cgst_pct": round(cgst_pct, 2),
                    "cgst_amount": round(cgst_amt, 2),
                    "sgst_pct": round(sgst_pct, 2),
                    "sgst_amount": round(sgst_amt, 2),
                    "igst_pct": round(igst_pct, 2),
                    "igst_amount": round(igst_amt, 2),
                    "total_tax": round(item.tax_amount, 2),
                    "total_invoice_value": round(inv.total_amount, 2),
                    "tds_pct": 0.0,
                    "tds_amount": 0.0,
                    "outstanding_amount": round(outstanding_amount, 2),
                    "remarks": inv.status
                })
                
        return rows

    @staticmethod
    def generate_sales_summary_excel(rows: List[Dict[str, Any]]):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Summary"

        ws.views.sheetView[0].showGridLines = True

        font_family = "Segoe UI"
        
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name=font_family, size=10)
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        totals_font = Font(name=font_family, size=11, bold=True)
        totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        thin_border_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        double_bottom_border = Border(
            left=thin_border_side, right=thin_border_side, 
            top=thin_border_side, 
            bottom=Side(border_style="double", color="1B4332")
        )

        headers = [
            "Invoice Number", "Invoice Date", "Payment Date", "Payment Mode", 
            "Customer Name", "Customer GSTIN", "Place of Supply", "HSN/SAC Code", 
            "Product/Service Description", "Taxable Value", "Discount", "CGST %", 
            "CGST Amount", "SGST %", "SGST Amount", "IGST %", "IGST Amount", 
            "Total Tax", "Total Invoice Value", "TDS %", "TDS Amount", 
            "Outstanding Amount", "Remarks"
        ]

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        col_alignments = {
            1: align_left,
            2: align_center,
            3: align_center,
            4: align_center,
            5: align_left,
            6: align_center,
            7: align_left,
            8: align_center,
            9: align_left,
            10: align_right,
            11: align_right,
            12: align_right,
            13: align_right,
            14: align_right,
            15: align_right,
            16: align_right,
            17: align_right,
            18: align_right,
            19: align_right,
            20: align_right,
            21: align_right,
            22: align_right,
            23: align_left
        }

        currency_format = '₹#,##,##0.00'
        percent_format = '0.00"%"'

        row_num = 2
        for data in rows:
            row_data = [
                data["invoice_number"],
                data["invoice_date"],
                data["payment_date"] if data["payment_date"] else "-",
                data["payment_mode"] if data["payment_mode"] else "-",
                data["customer_name"],
                data["customer_gstin"] if data["customer_gstin"] else "-",
                data["place_of_supply"],
                data["hsn_code"] if data["hsn_code"] else "-",
                data["product_description"],
                data["taxable_value"],
                data["discount"],
                data["cgst_pct"] / 100.0,
                data["cgst_amount"],
                data["sgst_pct"] / 100.0,
                data["sgst_amount"],
                data["igst_pct"] / 100.0,
                data["igst_amount"],
                data["total_tax"],
                data["total_invoice_value"],
                data["tds_pct"] / 100.0,
                data["tds_amount"],
                data["outstanding_amount"],
                data["remarks"]
            ]
            
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 20
            
            is_zebra = (row_num % 2 == 1)
            row_fill = zebra_fill if is_zebra else white_fill
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.alignment = col_alignments.get(col_idx, align_left)
                cell.border = thin_border
                
                if col_idx in [10, 11, 13, 15, 17, 18, 19, 21, 22]:
                    cell.number_format = currency_format
                elif col_idx in [12, 14, 16, 20]:
                    cell.number_format = percent_format
            
            row_num += 1

        ws.row_dimensions[row_num].height = 24
        
        total_label_cell = ws.cell(row=row_num, column=1, value="TOTAL")
        total_label_cell.font = totals_font
        total_label_cell.fill = totals_fill
        total_label_cell.alignment = align_center
        total_label_cell.border = double_bottom_border
        
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        for col_idx in range(2, 10):
            ws.cell(row=row_num, column=col_idx).fill = totals_fill
            ws.cell(row=row_num, column=col_idx).border = double_bottom_border

        for col_idx in range(10, 24):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.border = double_bottom_border
            
            col_letter = get_column_letter(col_idx)
            
            if col_idx in [10, 11, 13, 15, 17, 18, 19, 21, 22]:
                cell.value = f"=SUM({col_letter}2:{col_letter}{row_num - 1})"
                cell.number_format = currency_format
                cell.alignment = align_right
            elif col_idx in [12, 14, 16, 20]:
                cell.value = ""
                cell.alignment = align_center
            else:
                cell.value = ""
                cell.alignment = align_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "₹99,99,999.00"
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def get_trial_balance(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        company_id: Optional[UUID] = None
    ) -> TrialBalanceResponse:
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text
        from app.schemas.accounts import TrialBalanceRow, TrialBalanceTotals, TrialBalanceResponse

        if start_date is None or end_date is None:
            def_start, def_end = current_fy_dates()
            start_date = start_date or def_start
            end_date = end_date or def_end

        sql_query = text("""
            WITH journal_agg AS (
              SELECT
                jl.ledger_id,
                SUM(CASE WHEN jl.dr_cr='Dr' AND je.date < :start_date THEN jl.amount ELSE 0 END) AS pre_dr,
                SUM(CASE WHEN jl.dr_cr='Cr' AND je.date < :start_date THEN jl.amount ELSE 0 END) AS pre_cr,
                SUM(CASE WHEN jl.dr_cr='Dr' AND je.date BETWEEN :start_date AND :end_date THEN jl.amount ELSE 0 END) AS period_dr,
                SUM(CASE WHEN jl.dr_cr='Cr' AND je.date BETWEEN :start_date AND :end_date THEN jl.amount ELSE 0 END) AS period_cr
              FROM journal_entry_lines jl
              JOIN journal_entries je ON je.id = jl.journal_entry_id
              WHERE je.is_active = true AND je.is_reversed = false
                AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
              GROUP BY jl.ledger_id
            )
            SELECT 
                la.id AS ledger_id,
                la.code AS code,
                la.name AS name,
                ag.name AS group_name,
                la.opening_bal AS base_opening,
                la.opening_bal_type AS base_opening_type,
                COALESCE(ja.pre_dr, 0) AS pre_dr,
                COALESCE(ja.pre_cr, 0) AS pre_cr,
                COALESCE(ja.period_dr, 0) AS period_dr,
                COALESCE(ja.period_cr, 0) AS period_cr
            FROM ledger_accounts la
            JOIN account_groups ag ON la.group_id = ag.id
            LEFT JOIN journal_agg ja ON ja.ledger_id = la.id
            WHERE la.is_active = true
            ORDER BY la.code ASC
        """)

        result = await db.execute(sql_query, {"start_date": start_date, "end_date": end_date, "company_id": company_id})
        rows = result.all()

        tb_rows = []
        op_dr_tot = Decimal("0.00")
        op_cr_tot = Decimal("0.00")
        mov_dr_tot = Decimal("0.00")
        mov_cr_tot = Decimal("0.00")
        cl_dr_tot = Decimal("0.00")
        cl_cr_tot = Decimal("0.00")

        for row in rows:
            base_opening = Decimal(str(row.base_opening))
            pre_dr = Decimal(str(row.pre_dr))
            pre_cr = Decimal(str(row.pre_cr))
            period_dr = Decimal(str(row.period_dr))
            period_cr = Decimal(str(row.period_cr))

            if row.base_opening_type == "Dr":
                eff_open_dr = base_opening + pre_dr
                eff_open_cr = pre_cr
            else:
                eff_open_dr = pre_dr
                eff_open_cr = base_opening + pre_cr

            opening_net = eff_open_dr - eff_open_cr
            if opening_net >= 0:
                op_dr = opening_net
                op_cr = Decimal("0.00")
                op_type = "Dr"
            else:
                op_dr = Decimal("0.00")
                op_cr = -opening_net
                op_type = "Cr"

            closing_dr = op_dr + period_dr
            closing_cr = op_cr + period_cr
            closing_net = closing_dr - closing_cr

            if closing_net >= 0:
                cl_dr = closing_net
                cl_cr = Decimal("0.00")
                cl_type = "Dr"
            else:
                cl_dr = Decimal("0.00")
                cl_cr = -closing_net
                cl_type = "Cr"

            op_dr_tot += op_dr
            op_cr_tot += op_cr
            mov_dr_tot += period_dr
            mov_cr_tot += period_cr
            cl_dr_tot += cl_dr
            cl_cr_tot += cl_cr

            tb_rows.append(
                TrialBalanceRow(
                    ledger_id=row.ledger_id,
                    code=row.code,
                    name=row.name,
                    group_name=row.group_name,
                    opening_dr=op_dr,
                    opening_cr=op_cr,
                    opening_bal_type=op_type,
                    movement_dr=period_dr,
                    movement_cr=period_cr,
                    closing_dr=cl_dr,
                    closing_cr=cl_cr,
                    closing_bal_type=cl_type
                )
            )

        totals = TrialBalanceTotals(
            opening_dr_total=op_dr_tot,
            opening_cr_total=op_cr_tot,
            movement_dr_total=mov_dr_tot,
            movement_cr_total=mov_cr_tot,
            closing_dr_total=cl_dr_tot,
            closing_cr_total=cl_cr_tot
        )

        return TrialBalanceResponse(rows=tb_rows, totals=totals)

    @staticmethod
    async def get_general_ledger(
        db: AsyncSession,
        ledger_id: UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        skip: int = 0,
        limit: int = 100,
        company_id: Optional[UUID] = None
    ) -> GeneralLedgerResponse:
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text
        from fastapi import HTTPException
        from app.models.accounts import LedgerAccount, AccountGroup
        from app.schemas.accounts import GeneralLedgerLine, GeneralLedgerResponse

        if start_date is None or end_date is None:
            def_start, def_end = current_fy_dates()
            start_date = start_date or def_start
            end_date = end_date or def_end

        # 1. Fetch Ledger details
        q_ledger = await db.execute(
            select(LedgerAccount)
            .join(AccountGroup, LedgerAccount.group_id == AccountGroup.id)
            .filter(LedgerAccount.id == ledger_id)
            .options(selectinload(LedgerAccount.group))
        )
        ledger = q_ledger.scalar_one_or_none()
        if not ledger:
            raise HTTPException(status_code=404, detail="Ledger account not found.")

        # 2. Compute pre-period journal movements
        q_pre = await db.execute(
            text("""
                SELECT
                    SUM(CASE WHEN jl.dr_cr = 'Dr' THEN jl.amount ELSE 0 END) as pre_dr,
                    SUM(CASE WHEN jl.dr_cr = 'Cr' THEN jl.amount ELSE 0 END) as pre_cr
                FROM journal_entry_lines jl
                JOIN journal_entries je ON je.id = jl.journal_entry_id
                WHERE je.is_active = true AND je.is_reversed = false AND jl.ledger_id = :ledger_id AND je.date < :start_date
                  AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
            """),
            {"ledger_id": ledger_id, "start_date": start_date, "company_id": company_id}
        )
        pre_row = q_pre.one()
        pre_dr = Decimal(str(pre_row.pre_dr or 0.00))
        pre_cr = Decimal(str(pre_row.pre_cr or 0.00))

        # Effective Opening Balance
        base_bal = Decimal(str(ledger.opening_bal))
        if ledger.opening_bal_type == "Dr":
            signed_balance = base_bal + pre_dr - pre_cr
        else:
            signed_balance = pre_dr - pre_cr - base_bal

        opening_balance = abs(signed_balance)
        opening_balance_type = "Dr" if signed_balance >= 0 else "Cr"

        # 3. Fetch period transaction lines
        sql_lines = text("""
            WITH sibling_summary AS (
                SELECT
                    journal_entry_id,
                    COUNT(id) as line_count,
                    CAST(MAX(CASE WHEN ledger_id != :ledger_id THEN CAST(ledger_id AS TEXT) ELSE NULL END) AS UUID) as other_ledger_id
                FROM journal_entry_lines
                GROUP BY journal_entry_id
            )
            SELECT
                jl.journal_entry_id,
                je.date,
                vt.name as voucher_type,
                je.voucher_number,
                je.narration as entry_narration,
                jl.narration as line_narration,
                jl.dr_cr,
                jl.amount,
                ss.line_count,
                ola.name as other_ledger_name
            FROM journal_entry_lines jl
            JOIN journal_entries je ON je.id = jl.journal_entry_id
            JOIN voucher_types vt ON vt.id = je.voucher_type_id
            JOIN sibling_summary ss ON ss.journal_entry_id = jl.journal_entry_id
            LEFT JOIN ledger_accounts ola ON ola.id = ss.other_ledger_id
            WHERE je.is_active = true AND jl.ledger_id = :ledger_id AND je.date BETWEEN :start_date AND :end_date
              AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
            ORDER BY je.date ASC, je.created_at ASC, jl.id ASC
            OFFSET :skip LIMIT :limit
        """)

        res_lines = await db.execute(
            sql_lines,
            {
                "ledger_id": ledger_id,
                "start_date": start_date,
                "end_date": end_date,
                "company_id": company_id,
                "skip": skip,
                "limit": limit
            }
        )
        period_rows = res_lines.all()

        lines = []
        total_debit = Decimal("0.00")
        total_credit = Decimal("0.00")

        # Sum offset movements if skip > 0
        if skip > 0:
            q_offset_mov = await db.execute(
                text("""
                    SELECT 
                        SUM(CASE WHEN dr_cr = 'Dr' THEN amount ELSE 0 END) as offset_dr,
                        SUM(CASE WHEN dr_cr = 'Cr' THEN amount ELSE 0 END) as offset_cr
                    FROM (
                        SELECT jl.dr_cr, jl.amount
                        FROM journal_entry_lines jl
                        JOIN journal_entries je ON je.id = jl.journal_entry_id
                        WHERE je.is_active = true AND jl.ledger_id = :ledger_id AND je.date BETWEEN :start_date AND :end_date
                          AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
                        ORDER BY je.date ASC, je.created_at ASC, jl.id ASC
                        LIMIT :skip
                    ) as offset_rows
                """),
                {"ledger_id": ledger_id, "start_date": start_date, "end_date": end_date, "company_id": company_id, "skip": skip}
            )
            offset_row = q_offset_mov.one()
            offset_dr = Decimal(str(offset_row.offset_dr or 0.00))
            offset_cr = Decimal(str(offset_row.offset_cr or 0.00))
            signed_balance += offset_dr - offset_cr

        for row in period_rows:
            amt = Decimal(str(row.amount))
            dr_cr = row.dr_cr

            if dr_cr == "Dr":
                total_debit += amt
                signed_balance += amt
            else:
                total_credit += amt
                signed_balance -= amt

            # Resolve particulars
            line_count = row.line_count
            if line_count == 2:
                particulars = row.other_ledger_name or "N/A"
            elif row.line_narration:
                particulars = row.line_narration
            elif row.entry_narration:
                particulars = row.entry_narration
            else:
                particulars = f"Multiple accounts ({line_count} lines)"

            running_balance = abs(signed_balance)
            running_balance_type = "Dr" if signed_balance >= 0 else "Cr"

            lines.append(
                GeneralLedgerLine(
                    journal_entry_id=row.journal_entry_id,
                    date=row.date,
                    voucher_type=row.voucher_type,
                    reference_no=row.voucher_number or "N/A",
                    narration=row.line_narration or row.entry_narration,
                    particulars=particulars,
                    debit=amt if dr_cr == "Dr" else Decimal("0.00"),
                    credit=amt if dr_cr == "Cr" else Decimal("0.00"),
                    running_balance=running_balance,
                    running_balance_type=running_balance_type
                )
            )

        # Compute closing totals for the filtered period
        q_tot = await db.execute(
            text("""
                SELECT
                    SUM(CASE WHEN jl.dr_cr = 'Dr' THEN jl.amount ELSE 0 END) as tot_dr,
                    SUM(CASE WHEN jl.dr_cr = 'Cr' THEN jl.amount ELSE 0 END) as tot_cr
                FROM journal_entry_lines jl
                JOIN journal_entries je ON je.id = jl.journal_entry_id
                WHERE je.is_active = true AND jl.ledger_id = :ledger_id AND je.date BETWEEN :start_date AND :end_date
                  AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
            """),
            {"ledger_id": ledger_id, "start_date": start_date, "end_date": end_date, "company_id": company_id}
        )
        tot_row = q_tot.one()
        tot_dr = Decimal(str(tot_row.tot_dr or 0.00))
        tot_cr = Decimal(str(tot_row.tot_cr or 0.00))

        if ledger.opening_bal_type == "Dr":
            final_signed = base_bal + pre_dr - pre_cr + tot_dr - tot_cr
        else:
            final_signed = pre_dr - pre_cr - base_bal + tot_dr - tot_cr

        closing_balance = abs(final_signed)
        closing_balance_type = "Dr" if final_signed >= 0 else "Cr"

        return GeneralLedgerResponse(
            ledger_id=ledger.id,
            code=ledger.code,
            name=ledger.name,
            group_name=ledger.group.name,
            opening_balance=opening_balance,
            opening_balance_type=opening_balance_type,
            closing_balance=closing_balance,
            closing_balance_type=closing_balance_type,
            lines=lines,
            total_debit=tot_dr,
            total_credit=tot_cr
        )

    @staticmethod
    async def get_day_book(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        company_id: Optional[UUID] = None,
        skip: int = 0,
        limit: int = 100
    ) -> DayBookResponse:
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text
        from app.schemas.accounts import DayBookEntry, DayBookLine, DayBookResponse

        if start_date is None or end_date is None:
            def_start, def_end = current_fy_dates()
            start_date = start_date or def_start
            end_date = end_date or def_end

        sql_entries = text("""
            SELECT DISTINCT je.id, je.date, vt.name as voucher_type, je.voucher_number, je.narration, je.is_reversed, je.created_at
            FROM journal_entries je
            JOIN voucher_types vt ON vt.id = je.voucher_type_id
            WHERE je.is_active = true AND je.date BETWEEN :start_date AND :end_date
              AND (CAST(:company_id AS UUID) IS NULL OR je.company_id = :company_id)
            ORDER BY je.date ASC, je.created_at ASC, je.id ASC
            OFFSET :skip LIMIT :limit
        """)

        res_entries = await db.execute(
            sql_entries,
            {
                "start_date": start_date,
                "end_date": end_date,
                "company_id": company_id,
                "skip": skip,
                "limit": limit
            }
        )
        entries_rows = res_entries.all()

        entries = []
        total_debit = Decimal("0.00")
        total_credit = Decimal("0.00")

        for row in entries_rows:
            sql_lines = text("""
                SELECT la.code, la.name, jl.dr_cr, jl.amount, jl.narration
                FROM journal_entry_lines jl
                JOIN ledger_accounts la ON la.id = jl.ledger_id
                WHERE jl.journal_entry_id = :journal_entry_id
                ORDER BY jl.dr_cr DESC, la.code ASC
            """)
            res_lines = await db.execute(sql_lines, {"journal_entry_id": row.id})
            lines_rows = res_lines.all()

            db_lines = []
            for l_row in lines_rows:
                amt = Decimal(str(l_row.amount))
                if l_row.dr_cr == "Dr":
                    total_debit += amt
                else:
                    total_credit += amt

                db_lines.append(
                    DayBookLine(
                        ledger_code=l_row.code,
                        ledger_name=l_row.name,
                        dr_cr=l_row.dr_cr,
                        amount=amt,
                        narration=l_row.narration
                    )
                )

            entries.append(
                DayBookEntry(
                    journal_entry_id=row.id,
                    date=row.date,
                    voucher_type=row.voucher_type,
                    voucher_number=row.voucher_number or "N/A",
                    narration=row.narration,
                    is_reversed=row.is_reversed,
                    lines=db_lines
                )
            )

        return DayBookResponse(
            entries=entries,
            total_debit=total_debit,
            total_credit=total_credit
        )

    @staticmethod
    async def get_purchase_register(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        supplier_id: Optional[UUID] = None,
        company_id: Optional[UUID] = None,
        skip: int = 0,
        limit: int = 100
    ) -> PurchaseRegisterResponse:
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text
        from app.schemas.accounts import PurchaseRegisterRow, PurchaseRegisterTotals, PurchaseRegisterResponse

        if start_date is None or end_date is None:
            def_start, def_end = current_fy_dates()
            start_date = start_date or def_start
            end_date = end_date or def_end

        sql_rows = text("""
            SELECT 
                pe.id,
                pe.billing_date,
                pe.invoice_number,
                s.name as supplier_name,
                s.gstin as supplier_gstin,
                s.address as supplier_address,
                c.state_code as branch_state_code,
                c.gstin as branch_gstin,
                pe.subtotal,
                pe.cgst_amount,
                pe.sgst_amount,
                pe.igst_amount,
                pe.tax_amount,
                pe.total_amount,
                pe.status
            FROM purchase_entries pe
            JOIN suppliers s ON s.id = pe.supplier_id
            JOIN companies c ON c.id = pe.company_id
            WHERE pe.status != 'Cancelled' AND pe.billing_date BETWEEN :start_date AND :end_date
              AND (CAST(:supplier_id AS UUID) IS NULL OR pe.supplier_id = :supplier_id)
              AND (CAST(:company_id AS UUID) IS NULL OR pe.company_id = :company_id)
            ORDER BY pe.billing_date ASC, pe.invoice_number ASC, pe.id ASC
            OFFSET :skip LIMIT :limit
        """)

        res_rows = await db.execute(
            sql_rows,
            {
                "start_date": start_date,
                "end_date": end_date,
                "supplier_id": supplier_id,
                "company_id": company_id,
                "skip": skip,
                "limit": limit
            }
        )
        rows_all = res_rows.all()

        STATE_CODES = {
            "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh", "05": "Uttarakhand",
            "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar",
            "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
            "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
            "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat", "26": "Dadra and Nagar Haveli and Daman and Diu",
            "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
            "38": "Ladakh"
        }

        register_rows = []
        for r in rows_all:
            place_of_supply = "Other"
            gstin = r.supplier_gstin or ""
            state_code = gstin[:2] if gstin else None
            if state_code in STATE_CODES:
                place_of_supply = f"{STATE_CODES[state_code]} ({state_code})"
            elif r.supplier_address:
                place_of_supply = r.supplier_address.split(",")[-1].strip()

            register_rows.append(
                PurchaseRegisterRow(
                    purchase_entry_id=r.id,
                    billing_date=r.billing_date.date() if isinstance(r.billing_date, datetime) else r.billing_date,
                    invoice_number=r.invoice_number,
                    supplier_name=r.supplier_name,
                    supplier_gstin=gstin,
                    place_of_supply=place_of_supply,
                    taxable_value=Decimal(str(r.subtotal)),
                    cgst_amount=Decimal(str(r.cgst_amount)),
                    sgst_amount=Decimal(str(r.sgst_amount)),
                    igst_amount=Decimal(str(r.igst_amount)),
                    total_tax=Decimal(str(r.tax_amount)),
                    total_amount=Decimal(str(r.total_amount)),
                    status=r.status
                )
            )

        sql_totals = text("""
            SELECT 
                SUM(pe.subtotal) as sub_total,
                SUM(pe.cgst_amount) as cgst_total,
                SUM(pe.sgst_amount) as sgst_total,
                SUM(pe.igst_amount) as igst_total,
                SUM(pe.tax_amount) as tax_total,
                SUM(pe.total_amount) as grand_total
            FROM purchase_entries pe
            WHERE pe.status != 'Cancelled' AND pe.billing_date BETWEEN :start_date AND :end_date
              AND (CAST(:supplier_id AS UUID) IS NULL OR pe.supplier_id = :supplier_id)
              AND (CAST(:company_id AS UUID) IS NULL OR pe.company_id = :company_id)
        """)
        res_totals = await db.execute(
            sql_totals,
            {
                "start_date": start_date,
                "end_date": end_date,
                "supplier_id": supplier_id,
                "company_id": company_id
            }
        )
        totals_row = res_totals.one()
        totals = PurchaseRegisterTotals(
            taxable_value_total=Decimal(str(totals_row.sub_total or 0.00)),
            cgst_amount_total=Decimal(str(totals_row.cgst_total or 0.00)),
            sgst_amount_total=Decimal(str(totals_row.sgst_total or 0.00)),
            igst_amount_total=Decimal(str(totals_row.igst_total or 0.00)),
            total_tax_total=Decimal(str(totals_row.tax_total or 0.00)),
            total_amount_total=Decimal(str(totals_row.grand_total or 0.00))
        )

        return PurchaseRegisterResponse(rows=register_rows, totals=totals)

    @staticmethod
    async def get_sales_register(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        customer_id: Optional[UUID] = None,
        company_id: Optional[UUID] = None,
        skip: int = 0,
        limit: int = 100
    ) -> SalesRegisterResponse:
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text
        from app.schemas.accounts import SalesRegisterRow, SalesRegisterTotals, SalesRegisterResponse

        if start_date is None or end_date is None:
            def_start, def_end = current_fy_dates()
            start_date = start_date or def_start
            end_date = end_date or def_end

        sql_rows = text("""
            SELECT 
                inv.id,
                inv.date as invoice_date,
                inv.invoice_number,
                cust.name as customer_name,
                cust.gstin as customer_gstin,
                cust.billing_address as customer_address,
                comp.state_code as branch_state_code,
                comp.gstin as branch_gstin,
                inv.subtotal,
                inv.tax_amount,
                inv.total_amount,
                inv.status,
                inv.gst_breakup
            FROM invoices inv
            JOIN sales_orders so ON so.id = inv.sales_order_id
            JOIN customers cust ON cust.id = so.customer_id
            JOIN companies comp ON comp.id = inv.company_id
            WHERE inv.status != 'Cancelled' AND inv.date BETWEEN :start_date AND :end_date
              AND (CAST(:customer_id AS UUID) IS NULL OR so.customer_id = :customer_id)
              AND (CAST(:company_id AS UUID) IS NULL OR inv.company_id = :company_id)
            ORDER BY inv.date ASC, inv.invoice_number ASC, inv.id ASC
            OFFSET :skip LIMIT :limit
        """)

        res_rows = await db.execute(
            sql_rows,
            {
                "start_date": start_date,
                "end_date": end_date,
                "customer_id": customer_id,
                "company_id": company_id,
                "skip": skip,
                "limit": limit
            }
        )
        rows_all = res_rows.all()

        STATE_CODES = {
            "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh", "05": "Uttarakhand",
            "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar",
            "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
            "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
            "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat", "26": "Dadra and Nagar Haveli and Daman and Diu",
            "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
            "38": "Ladakh"
        }

        register_rows = []
        for r in rows_all:
            place_of_supply = "Other"
            gstin = r.customer_gstin or ""
            state_code = gstin[:2] if gstin else None
            if state_code in STATE_CODES:
                place_of_supply = f"{STATE_CODES[state_code]} ({state_code})"
            elif r.customer_address:
                place_of_supply = r.customer_address.split(",")[-1].strip()

            import json
            gst_breakup = r.gst_breakup
            if isinstance(gst_breakup, str):
                try:
                    gst_breakup = json.loads(gst_breakup)
                except Exception:
                    gst_breakup = {}
            elif not isinstance(gst_breakup, dict):
                gst_breakup = {}

            cgst_amount = Decimal(str(gst_breakup.get("cgst", 0.00)))
            sgst_amount = Decimal(str(gst_breakup.get("sgst", 0.00)))
            igst_amount = Decimal(str(gst_breakup.get("igst", 0.00)))

            register_rows.append(
                SalesRegisterRow(
                    invoice_id=r.id,
                    invoice_date=r.invoice_date.date() if isinstance(r.invoice_date, datetime) else r.invoice_date,
                    invoice_number=r.invoice_number,
                    customer_name=r.customer_name,
                    customer_gstin=gstin,
                    place_of_supply=place_of_supply,
                    taxable_value=Decimal(str(r.subtotal)),
                    cgst_amount=cgst_amount,
                    sgst_amount=sgst_amount,
                    igst_amount=igst_amount,
                    total_tax=Decimal(str(r.tax_amount)),
                    total_amount=Decimal(str(r.total_amount)),
                    status=r.status
                )
            )

        sql_totals = text("""
            SELECT 
                SUM(inv.subtotal) as sub_total,
                SUM(inv.tax_amount) as tax_total,
                SUM(inv.total_amount) as grand_total
            FROM invoices inv
            JOIN sales_orders so ON so.id = inv.sales_order_id
            WHERE inv.status != 'Cancelled' AND inv.date BETWEEN :start_date AND :end_date
              AND (CAST(:customer_id AS UUID) IS NULL OR so.customer_id = :customer_id)
              AND (CAST(:company_id AS UUID) IS NULL OR inv.company_id = :company_id)
        """)
        res_totals = await db.execute(
            sql_totals,
            {
                "start_date": start_date,
                "end_date": end_date,
                "customer_id": customer_id,
                "company_id": company_id
            }
        )
        totals_row = res_totals.one()

        sql_tax_totals = text("""
            SELECT 
                SUM(COALESCE(CAST(inv.gst_breakup->>'cgst' AS NUMERIC), 0)) as cgst_total,
                SUM(COALESCE(CAST(inv.gst_breakup->>'sgst' AS NUMERIC), 0)) as sgst_total,
                SUM(COALESCE(CAST(inv.gst_breakup->>'igst' AS NUMERIC), 0)) as igst_total
            FROM invoices inv
            JOIN sales_orders so ON so.id = inv.sales_order_id
            WHERE inv.status != 'Cancelled' AND inv.date BETWEEN :start_date AND :end_date
              AND (CAST(:customer_id AS UUID) IS NULL OR so.customer_id = :customer_id)
              AND (CAST(:company_id AS UUID) IS NULL OR inv.company_id = :company_id)
        """)
        res_tax = await db.execute(
            sql_tax_totals,
            {
                "start_date": start_date,
                "end_date": end_date,
                "customer_id": customer_id,
                "company_id": company_id
            }
        )
        tax_row = res_tax.one()

        totals = SalesRegisterTotals(
            taxable_value_total=Decimal(str(totals_row.sub_total or 0.00)),
            cgst_amount_total=Decimal(str(tax_row.cgst_total or 0.00)),
            sgst_amount_total=Decimal(str(tax_row.sgst_total or 0.00)),
            igst_amount_total=Decimal(str(tax_row.igst_total or 0.00)),
            total_tax_total=Decimal(str(totals_row.tax_total or 0.00)),
            total_amount_total=Decimal(str(totals_row.grand_total or 0.00))
        )

        return SalesRegisterResponse(rows=register_rows, totals=totals)

    @staticmethod
    async def get_cash_book(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        skip: int = 0,
        limit: int = 100
    ) -> Dict[str, Any]:
        """
        Retrieves a Cash Book showing transaction lines for Cash ledgers.
        """
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text

        if start_date is None or end_date is None:
            start_date, end_date = current_fy_dates()

        prev_day = start_date - timedelta(days=1)

        # 1. Get all Cash Group Ledger IDs
        ledgers_sql = text("""
            SELECT id FROM ledger_accounts la
            WHERE la.is_active = true 
              AND (la.group_id = '12300000-0000-0000-0000-000000000001' 
                OR la.group_id IN (SELECT id FROM account_groups WHERE parent_id = '12300000-0000-0000-0000-000000000001')
                OR la.name ILIKE '%cash%')
        """)
        ledger_res = await db.execute(ledgers_sql)
        ledger_ids = [row.id for row in ledger_res.all()]

        if not ledger_ids:
            return {
                "opening_balance": 0.0,
                "closing_balance": 0.0,
                "total_debit": 0.0,
                "total_credit": 0.0,
                "records": []
            }

        # 2. Compute opening balance
        op_sql = text("""
            SELECT
              SUM(
                (CASE WHEN la.opening_bal_type = 'Dr' THEN la.opening_bal ELSE -la.opening_bal END) +
                COALESCE(ja.total_dr, 0) - COALESCE(ja.total_cr, 0)
              ) AS opening_bal
            FROM ledger_accounts la
            LEFT JOIN (
              SELECT
                jl.ledger_id,
                SUM(CASE WHEN jl.dr_cr = 'Dr' THEN jl.amount ELSE 0 END) AS total_dr,
                SUM(CASE WHEN jl.dr_cr = 'Cr' THEN jl.amount ELSE 0 END) AS total_cr
              FROM journal_entry_lines jl
              JOIN journal_entries je ON je.id = jl.journal_entry_id
              WHERE je.is_active = true AND je.is_reversed = false AND je.date <= :prev_day
              GROUP BY jl.ledger_id
            ) ja ON ja.ledger_id = la.id
            WHERE la.id = ANY(:ledger_ids)
        """)
        op_res = await db.execute(op_sql, {"prev_day": prev_day, "ledger_ids": ledger_ids})
        opening_balance = Decimal(str(op_res.scalar() or 0.00))

        # 3. Fetch transaction lines within date range
        tx_sql = text("""
            SELECT
              je.date AS date,
              je.voucher_number AS voucher_number,
              jl.dr_cr AS dr_cr,
              jl.amount AS amount,
              jl.narration AS narration,
              la.name AS ledger_name,
              (
                SELECT STRING_AGG(comp_la.name, ', ')
                FROM journal_entry_lines comp_jl
                JOIN ledger_accounts comp_la ON comp_la.id = comp_jl.ledger_id
                WHERE comp_jl.journal_entry_id = je.id AND comp_jl.ledger_id != jl.ledger_id
              ) AS particulars
            FROM journal_entry_lines jl
            JOIN journal_entries je ON je.id = jl.journal_entry_id
            JOIN ledger_accounts la ON la.id = jl.ledger_id
            WHERE je.is_active = true AND je.is_reversed = false 
              AND je.date BETWEEN :start_date AND :end_date
              AND jl.ledger_id = ANY(:ledger_ids)
            ORDER BY je.date ASC, je.voucher_number ASC
        """)
        tx_res = await db.execute(tx_sql, {
            "start_date": start_date,
            "end_date": end_date,
            "ledger_ids": ledger_ids
        })
        tx_rows = tx_res.all()

        records = []
        running_balance = opening_balance
        total_debit = Decimal("0.00")
        total_credit = Decimal("0.00")

        for r in tx_rows:
            amt = Decimal(str(r.amount))
            is_dr = r.dr_cr == 'Dr'
            if is_dr:
                debit = amt
                credit = Decimal("0.00")
                running_balance += amt
                total_debit += amt
            else:
                debit = Decimal("0.00")
                credit = amt
                running_balance -= amt
                total_credit += amt

            records.append({
                "date": r.date.isoformat() if isinstance(r.date, (date, datetime)) else str(r.date),
                "voucher_number": r.voucher_number or "JV-N/A",
                "particulars": r.particulars or "Multiple Ledgers",
                "ledger_name": r.ledger_name,
                "narration": r.narration,
                "debit": float(debit),
                "credit": float(credit),
                "running_balance": float(running_balance)
            })

        paginated_records = records[skip: skip + limit]

        return {
            "opening_balance": float(opening_balance),
            "closing_balance": float(running_balance),
            "total_debit": float(total_debit),
            "total_credit": float(total_credit),
            "total_count": len(records),
            "records": paginated_records
        }

    @staticmethod
    async def get_bank_book(
        db: AsyncSession,
        bank_ledger_id: Optional[UUID] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        skip: int = 0,
        limit: int = 100
    ) -> Dict[str, Any]:
        """
        Retrieves a Bank Book for a specific bank ledger or all bank ledgers.
        """
        from app.core.account_constants import current_fy_dates
        from decimal import Decimal
        from sqlalchemy import text

        if start_date is None or end_date is None:
            start_date, end_date = current_fy_dates()

        prev_day = start_date - timedelta(days=1)

        # 1. Identify bank ledger IDs
        if bank_ledger_id:
            ledger_ids = [bank_ledger_id]
        else:
            ledgers_sql = text("""
                SELECT id FROM ledger_accounts la
                WHERE la.is_active = true 
                  AND (la.name ILIKE '%bank%' OR la.name ILIKE '%saving%' OR la.name ILIKE '%current a/c%')
            """)
            ledger_res = await db.execute(ledgers_sql)
            ledger_ids = [row.id for row in ledger_res.all()]

        if not ledger_ids:
            return {
                "opening_balance": 0.0,
                "closing_balance": 0.0,
                "total_debit": 0.0,
                "total_credit": 0.0,
                "records": []
            }

        # 2. Compute opening balance
        op_sql = text("""
            SELECT
              SUM(
                (CASE WHEN la.opening_bal_type = 'Dr' THEN la.opening_bal ELSE -la.opening_bal END) +
                COALESCE(ja.total_dr, 0) - COALESCE(ja.total_cr, 0)
              ) AS opening_bal
            FROM ledger_accounts la
            LEFT JOIN (
              SELECT
                jl.ledger_id,
                SUM(CASE WHEN jl.dr_cr = 'Dr' THEN jl.amount ELSE 0 END) AS total_dr,
                SUM(CASE WHEN jl.dr_cr = 'Cr' THEN jl.amount ELSE 0 END) AS total_cr
              FROM journal_entry_lines jl
              JOIN journal_entries je ON je.id = jl.journal_entry_id
              WHERE je.is_active = true AND je.is_reversed = false AND je.date <= :prev_day
              GROUP BY jl.ledger_id
            ) ja ON ja.ledger_id = la.id
            WHERE la.id = ANY(:ledger_ids)
        """)
        op_res = await db.execute(op_sql, {"prev_day": prev_day, "ledger_ids": ledger_ids})
        opening_balance = Decimal(str(op_res.scalar() or 0.00))

        # 3. Fetch transaction lines within date range
        tx_sql = text("""
            SELECT
              je.date AS date,
              je.voucher_number AS voucher_number,
              jl.dr_cr AS dr_cr,
              jl.amount AS amount,
              jl.narration AS narration,
              la.name AS ledger_name,
              (
                SELECT STRING_AGG(comp_la.name, ', ')
                FROM journal_entry_lines comp_jl
                JOIN ledger_accounts comp_la ON comp_la.id = comp_jl.ledger_id
                WHERE comp_jl.journal_entry_id = je.id AND comp_jl.ledger_id != jl.ledger_id
              ) AS particulars
            FROM journal_entry_lines jl
            JOIN journal_entries je ON je.id = jl.journal_entry_id
            JOIN ledger_accounts la ON la.id = jl.ledger_id
            WHERE je.is_active = true AND je.is_reversed = false 
              AND je.date BETWEEN :start_date AND :end_date
              AND jl.ledger_id = ANY(:ledger_ids)
            ORDER BY je.date ASC, je.voucher_number ASC
        """)
        tx_res = await db.execute(tx_sql, {
            "start_date": start_date,
            "end_date": end_date,
            "ledger_ids": ledger_ids
        })
        tx_rows = tx_res.all()

        records = []
        running_balance = opening_balance
        total_debit = Decimal("0.00")
        total_credit = Decimal("0.00")

        for r in tx_rows:
            amt = Decimal(str(r.amount))
            is_dr = r.dr_cr == 'Dr'
            if is_dr:
                debit = amt
                credit = Decimal("0.00")
                running_balance += amt
                total_debit += amt
            else:
                debit = Decimal("0.00")
                credit = amt
                running_balance -= amt
                total_credit += amt

            records.append({
                "date": r.date.isoformat() if isinstance(r.date, (date, datetime)) else str(r.date),
                "voucher_number": r.voucher_number or "JV-N/A",
                "particulars": r.particulars or "Multiple Ledgers",
                "ledger_name": r.ledger_name,
                "narration": r.narration,
                "debit": float(debit),
                "credit": float(credit),
                "running_balance": float(running_balance)
            })

        paginated_records = records[skip: skip + limit]

        return {
            "opening_balance": float(opening_balance),
            "closing_balance": float(running_balance),
            "total_debit": float(total_debit),
            "total_credit": float(total_credit),
            "total_count": len(records),
            "records": paginated_records
        }

    @staticmethod
    async def get_journal_register(
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        skip: int = 0,
        limit: int = 100
    ) -> Dict[str, Any]:
        """
        Retrieves a Journal Register listing all posted journal entries with their lines.
        """
        from app.core.account_constants import current_fy_dates
        from sqlalchemy import text

        if start_date is None or end_date is None:
            start_date, end_date = current_fy_dates()

        # Count total
        count_sql = text("""
            SELECT COUNT(je.id) FROM journal_entries je
            WHERE je.is_active = true AND je.is_reversed = false
              AND je.date BETWEEN :start_date AND :end_date
        """)
        count_res = await db.execute(count_sql, {"start_date": start_date, "end_date": end_date})
        total = count_res.scalar() or 0

        # Fetch journal entries with limit
        je_sql = text("""
            SELECT 
              je.id AS entry_id,
              je.date AS date,
              je.voucher_number AS voucher_number,
              je.narration AS narration,
              vt.name AS voucher_type
            FROM journal_entries je
            JOIN voucher_types vt ON vt.id = je.voucher_type_id
            WHERE je.is_active = true AND je.is_reversed = false
              AND je.date BETWEEN :start_date AND :end_date
            ORDER BY je.date DESC, je.voucher_number DESC
            OFFSET :skip LIMIT :limit
        """)
        je_res = await db.execute(je_sql, {
            "start_date": start_date,
            "end_date": end_date,
            "skip": skip,
            "limit": limit
        })
        entries = je_res.all()

        if not entries:
            return {
                "total": total,
                "skip": skip,
                "limit": limit,
                "records": []
            }

        entry_ids = [str(e.entry_id) for e in entries]

        # Fetch lines for these entries
        lines_sql = text("""
            SELECT
              jl.journal_entry_id AS entry_id,
              jl.dr_cr AS dr_cr,
              jl.amount AS amount,
              jl.narration AS line_narration,
              la.name AS ledger_name,
              la.code AS ledger_code
            FROM journal_entry_lines jl
            JOIN ledger_accounts la ON la.id = jl.ledger_id
            WHERE jl.journal_entry_id = ANY(:entry_ids)
            ORDER BY jl.dr_cr DESC, la.name ASC
        """)
        lines_res = await db.execute(lines_sql, {"entry_ids": entry_ids})
        lines = lines_res.all()

        lines_by_entry = {}
        for l in lines:
            lines_by_entry.setdefault(str(l.entry_id), []).append({
                "ledger_code": l.ledger_code,
                "ledger_name": l.ledger_name,
                "dr_cr": l.dr_cr,
                "amount": float(l.amount),
                "narration": l.line_narration
            })

        records = []
        for e in entries:
            records.append({
                "id": str(e.entry_id),
                "date": e.date.isoformat() if isinstance(e.date, (date, datetime)) else str(e.date),
                "voucher_number": e.voucher_number or "N/A",
                "voucher_type": e.voucher_type,
                "narration": e.narration,
                "lines": lines_by_entry.get(str(e.entry_id), [])
            })

        return {
            "total": total,
            "skip": skip,
            "limit": limit,
            "records": records
        }

