from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from fastapi.responses import StreamingResponse
from datetime import date, datetime
from decimal import Decimal

def build_xlsx_response(
    headers: list[str],
    rows: list[list],
    sheet_name: str,
    report_name: str,
    start_date: date,
    end_date: date
) -> StreamingResponse:
    """Centralized utility to stream formatted Excel reports to the client."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True

    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name=font_family, size=10)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Write headers
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Format mappings
    CURRENCY_FORMAT = '₹#,##0.00'
    DATE_FORMAT = 'DD-MM-YYYY'

    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    row_num = 2
    for r in rows:
        ws.append(r)
        ws.row_dimensions[row_num].height = 20
        is_zebra = (row_num % 2 == 1)
        row_fill = zebra_fill if is_zebra else white_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border

            val = cell.value

            # Format cell values based on type
            if isinstance(val, (date, datetime)):
                cell.number_format = DATE_FORMAT
                cell.alignment = align_center
            elif isinstance(val, (int, float, Decimal)) and not isinstance(val, bool):
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = align_right
                cell.value = float(val) if isinstance(val, Decimal) else val
            elif isinstance(val, bool):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        row_num += 1

    # Freeze header and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths based on headers and content
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        header_text = headers[col_idx - 1]
        
        # Set default widths
        h_lower = header_text.lower()
        if "date" in h_lower:
            width = 12
        elif "narration" in h_lower or "particulars" in h_lower or "description" in h_lower or "name" in h_lower:
            width = 30
        elif "amount" in h_lower or "value" in h_lower or "debit" in h_lower or "credit" in h_lower or "total" in h_lower or "balance" in h_lower:
            width = 15
        else:
            width = max(len(header_text) + 4, 12)

        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_name}_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
